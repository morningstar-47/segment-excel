import argparse
import os
import sys
import logging
from typing import List, Tuple, Optional, Any
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook


def setup_logging(verbose: bool = False) -> None:
    """Configure le système de journalisation.
    
    Args:
        verbose: Si True, affiche les messages de debug
    """
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format='%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )


def parse_arguments() -> argparse.Namespace:
    """Parse les arguments de la ligne de commande.
    
    Returns:
        Les arguments analysés
    
    Raises:
        SystemExit: Si les arguments sont invalides
    """
    parser = argparse.ArgumentParser(
        description='Séparer les données de fichiers Excel en se basant sur des critères de filtrage'
    )
    parser.add_argument('path', type=str, help='Chemin d\'accès au fichier Excel')
    
    # Groupe mutuellement exclusif pour les options de filtrage de téléphone
    phone_group = parser.add_mutually_exclusive_group(required=True)
    phone_group.add_argument('-f', '--fixe', action='store_true', help='Filtrer sur Téléphone fixe')
    phone_group.add_argument('-m', '--mobile', action='store_true', help='Filtrer sur Numéro de téléphone')
    
    parser.add_argument('size', type=int, help='Définir la taille de ligne de sortie (max lignes par fichier)')
    parser.add_argument('-o', '--output', type=str, default='output', 
                        help='Dossier de sortie pour les fichiers générés (défaut: "output")')
    parser.add_argument('-v', '--verbose', action='store_true', help='Afficher les messages de débogage')
    
    args = parser.parse_args()
    
    # Validation supplémentaire
    if args.size <= 0:
        parser.error("La taille doit être un nombre positif")
        
    return args


def validate_input_file(file_path: str) -> None:
    """Vérifie si le fichier d'entrée existe et est valide.
    
    Args:
        file_path: Chemin vers le fichier Excel
    
    Raises:
        FileNotFoundError: Si le fichier n'existe pas
        ValueError: Si le fichier n'est pas un fichier Excel valide
    """
    path = Path(file_path)
    
    if not path.exists():
        raise FileNotFoundError(f"Le fichier '{file_path}' n'existe pas.")
    
    if not path.is_file():
        raise ValueError(f"'{file_path}' n'est pas un fichier.")
    
    if path.suffix.lower() not in ['.xlsx', '.xls']:
        raise ValueError(f"'{file_path}' n'est pas un fichier Excel (.xlsx ou .xls).")


def load_workbook(file_path: str) -> Workbook:
    """Charge un fichier Excel.
    
    Args:
        file_path: Chemin vers le fichier Excel
        
    Returns:
        Le classeur Excel chargé
        
    Raises:
        Exception: Si le chargement échoue
    """
    try:
        return openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        raise Exception(f"Erreur lors du chargement du fichier Excel: {str(e)}")


def filter_data(worksheet: Worksheet, column_name: str) -> Tuple[List[Any], List[List[Any]]]:
    """Filtre les données d'une feuille Excel.
    
    Args:
        worksheet: La feuille Excel à filtrer
        column_name: Le nom de la colonne à utiliser pour le filtrage
        
    Returns:
        Tuple contenant l'en-tête et les données filtrées
        
    Raises:
        ValueError: Si la colonne spécifiée n'est pas trouvée
    """
    # Récupérer l'en-tête
    headers = [cell.value for cell in next(worksheet.rows)]
    
    if column_name not in headers:
        raise ValueError(f"'{column_name}' n'est pas présent dans la liste des entêtes.")
    
    column_index = headers.index(column_name)
    
    # Filtrer les données
    filtered_data = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[column_index]:  # Vérifie si la valeur n'est pas None ou vide
            filtered_data.append(row)
    
    return headers, filtered_data


def split_data(headers: List[Any], data: List[List[Any]], chunk_size: int) -> List[List[List[Any]]]:
    """Divise les données en blocs de taille spécifiée.
    
    Args:
        headers: L'en-tête des données
        data: Les données à diviser
        chunk_size: Taille maximale de chaque bloc (sans compter l'en-tête)
        
    Returns:
        Liste de blocs, chacun contenant l'en-tête et les données
    """
    chunks = []
    current_chunk = [headers]
    
    for row in data:
        current_chunk.append(row)
        if len(current_chunk) > chunk_size:  # > car on compte aussi l'en-tête
            chunks.append(current_chunk)
            current_chunk = [headers]
    
    # Ajouter le dernier bloc s'il contient des données
    if len(current_chunk) > 1:  # > 1 car on veut au moins une ligne de données en plus de l'en-tête
        chunks.append(current_chunk)
    
    return chunks


def ensure_directory(directory_path: str) -> None:
    """Crée un répertoire s'il n'existe pas.
    
    Args:
        directory_path: Chemin du répertoire
    """
    path = Path(directory_path)
    path.mkdir(parents=True, exist_ok=True)


def save_chunks(chunks: List[List[List[Any]]], sheet_name: str, output_dir: str) -> None:
    """Sauvegarde les blocs de données dans des fichiers Excel séparés.
    
    Args:
        chunks: Liste des blocs de données
        sheet_name: Nom de la feuille d'origine
        output_dir: Répertoire de sortie
    """
    # Créer le dossier de sortie spécifique à la feuille
    sheet_folder_name = sheet_name.replace(" ", "_")
    sheet_output_dir = os.path.join(output_dir, sheet_folder_name)
    ensure_directory(sheet_output_dir)
    
    # Sauvegarder chaque bloc dans un fichier séparé
    for i, chunk in enumerate(chunks):
        output_filename = os.path.join(
            sheet_output_dir, 
            f"{sheet_name.capitalize()}_{i+1}.xlsx"
        )
        
        # Créer un nouveau fichier Excel
        new_workbook = openpyxl.Workbook()
        sheet = new_workbook.active
        
        # Ajouter les données
        for row in chunk:
            sheet.append(row)
        
        # Sauvegarder le fichier
        try:
            new_workbook.save(output_filename)
            logging.info(f"Fichier sauvegardé: {output_filename}")
        except Exception as e:
            logging.error(f"Erreur lors de la sauvegarde de {output_filename}: {str(e)}")


def process_worksheet(worksheet: Worksheet, phone_type: str, chunk_size: int, output_dir: str) -> int:
    """Traite une feuille Excel.
    
    Args:
        worksheet: La feuille Excel à traiter
        phone_type: Le type de téléphone à filtrer
        chunk_size: Taille maximale de chaque bloc
        output_dir: Répertoire de sortie
        
    Returns:
        Le nombre de fichiers générés
        
    Raises:
        ValueError: Si la colonne spécifiée n'est pas trouvée
    """
    try:
        logging.info(f"Traitement de la feuille '{worksheet.title}'...")
        
        # Filtrer les données
        headers, filtered_data = filter_data(worksheet, phone_type)
        logging.info(f"Données filtrées: {len(filtered_data)} lignes trouvées")
        
        # Diviser les données en blocs
        chunks = split_data(headers, filtered_data, chunk_size)
        logging.info(f"Données divisées en {len(chunks)} blocs")
        
        # Sauvegarder les blocs
        save_chunks(chunks, worksheet.title, output_dir)
        
        return len(chunks)
    except Exception as e:
        logging.error(f"Erreur lors du traitement de la feuille '{worksheet.title}': {str(e)}")
        return 0


def main() -> None:
    """Fonction principale du script."""
    try:
        # Analyser les arguments
        args = parse_arguments()
        
        # Configurer la journalisation
        setup_logging(args.verbose)
        
        # Déterminer le type de téléphone à filtrer
        phone_type = "Téléphone fixe" if args.fixe else "Numéro de téléphone"
        
        # Valider le fichier d'entrée
        validate_input_file(args.path)
        
        # Charger le fichier Excel
        logging.info(f"Chargement du fichier '{args.path}'...")
        workbook = load_workbook(args.path)
        
        # Créer le répertoire de sortie principal
        ensure_directory(args.output)
        
        # Traiter chaque feuille du classeur
        total_files = 0
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            files_generated = process_worksheet(sheet, phone_type, args.size, args.output)
            total_files += files_generated
        
        logging.info(f"Traitement terminé. {total_files} fichiers générés.")
        
    except Exception as e:
        logging.error(f"Erreur: {str(e)}")
        sys.exit(1)


if __name__ == '__main__':
    main()
